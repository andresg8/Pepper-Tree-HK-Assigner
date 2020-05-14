from datetime import *
from tkinter import *
from random import *
from hkGUI import *
from loading import *
from openpyxl import *
from openpyxl.styles import *
from threading import *
import tika
from tika import parser
import pickle
import os
import sys
dtoday = datetime.today()
today = date(dtoday.year, dtoday.month, dtoday.day)
wday = dtoday.weekday()

class HouseKeeper:
    def __init__(self, name, prevWorkDate, priority):
        self.name = name
        self.prevWorkDate = prevWorkDate
        self.priority = priority
        

class AllHouseKeepers:
    def __init__(self, housekeepers):
        self.housekeepers = housekeepers

    def addHouseKeeper(self, housekeeper):
        self.housekeepers.append(housekeeper)


class Room:
    def __init__(self, roomNumber, roomType, cleanType = "C/O", arrival = "", longStay = "", checkInDate = "", checkFT = ""):
        self.roomNumber = roomNumber
        self.roomType = roomType
        self.cleanType = cleanType
        self.arrival = arrival
        self.longStay = longStay
        self.checkInDate = checkInDate
        self.checkFT = checkFT

    def checkBlueness(self):
        '''
        Does a manual check on whether or not the room needs the sheets changed
        based on the check-in date compared to today's date. 
        '''
        date_split = self.checkInDate.split("/") # [MONTH, DAY, YEAR]
        try: CIDate = date(int(date_split[2]), int(date_split[0]), int(date_split[1]))
        except: print(date_split) 
        dif = today - CIDate
        if self.cleanType == "Stay" and not dif.days % 4 and dif.days > 1:
            ###print("FOUND ONE:", self.roomNumber)
            self.longStay = True

    def checkOrangeness(self):
        '''
        Does a manual check on whether or not the room needs the sheets changed
        based on the check-in date compared to today's date. 
        '''
        date_split = self.checkInDate.split("/") # [MONTH, DAY, YEAR]
        try: CIDate = date(int(date_split[2]), int(date_split[0]), int(date_split[1]))
        except: print(date_split) 
        dif = today - CIDate
        if self.cleanType == "Stay" and dif.days % 4 == 1 and dif.days > 1:
            ###print("FOUND ONE:", self.roomNumber)
            self.checkFT = True

    def __str__(self):
        strRep = ""
        if self.roomType in ("CK", "ACK", "KING"): strRep += "1 "
        if self.roomType in ("CQ", "ACQ", "QQ"): strRep += "2 "
        strRep += str(self.roomNumber) + " "
        if (self.arrival and self.cleanType == "C/O") or (self.longStay and self.cleanType == "Stay"): strRep += "#"
        strRep += str(self.cleanType)
        if (self.arrival and self.cleanType == "C/O") or (self.longStay and self.cleanType == "Stay"): strRep += "#"
        return strRep
    

class Rig():
    def __init__(self):
        self.stop = True
        
def outdatedCARError():
    Jerry = Rig()
    window = Tk()
    w = 275
    h = 250
    ws = window.winfo_screenwidth() 
    hs = window.winfo_screenheight() 
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    window.geometry('%dx%d+%d+%d' % (w, h, x, y))
    def onExitStop():
        Jerry.stop = True
        window.destroy()
    def onExitCont():
        Jerry.stop = False
        window.destroy()
    window.configure(background = "cornflower blue")
    msg = Label(window, text = "The Custom Associates Report you're using seems to be outdated" +
                " or it's the afternoon and you shouldn't be using an afternoon report"+
                ". It's strongly advised you cancel and obtain a current/reasonable"+
                " version of the report.", font = ("times", 14), wraplength = w,
                bg = "cornflower blue")
    cont = Button(window, text = "Continue\nAnyway", font = ("Bold"), width = 9, height = 3,
                  command = onExitCont, bg = "brown3")
    cancel = Button(window, text = "Cancel", font = ("Bold"), width = 9, height = 3,
                    command = onExitStop, bg = "chartreuse3")
    msg.pack(side = TOP)
    cont.pack(side = LEFT)
    cancel.pack(side = RIGHT)
    window.mainloop()
    return Jerry.stop

def generalError(eMsg):
    window = Tk()
    w = 275
    h = 250
    ws = window.winfo_screenwidth() 
    hs = window.winfo_screenheight() 
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    window.geometry('%dx%d+%d+%d' % (w, h, x, y))
    def onExitStop():
        window.destroy()
    window.configure(background = "light cyan")
    msg = Label(window, text = eMsg, font = ("times", 14), wraplength = w,
                bg = "light cyan")
    cancel = Button(window, text = "Exit", font = ("Bold"), width = 9, height = 3,
                    command = onExitStop, bg = "pale green")
    msg.pack(side = TOP)
    cancel.pack(side = BOTTOM)
    window.mainloop()
    return None

def goHome():
    CAR = "Custom Associates Report.pdf"
    true_cwd = os.getcwd()
    os.chdir("../")
    if os.getcwd().endswith("HK Assigner") or CAR in os.listdir(os.getcwd()):
        return
    else:
        os.chdir(true_cwd)
        print("Filepath Traversal Error Occured")

def smoothDivision(divisor : int, dividend : int) -> [int]:
    '''
    Handles splitting a dividend by a divisor as evenly as possible
    storage is populated by the resulting ints (within +-1 of each other)
    '''
    storage = []
    while divisor:
        quotient = divisor//dividend
        storage.append(quotient)
        divisor -= quotient
        dividend -= 1
    return storage


def queuePrintJob(pageNum, ws, wb):
    ws["Q1"] = pageNum
    wb.save('../Pre-Room Check List' + str(pageNum) + '.xlsx')
    wb.close()
    wb = load_workbook('Pre-Room Check List.xlsx')
    ws = wb["Sheet1"]

if __name__ == "__main__":
    try:
        AHK = pickle.load(open("housekeepers.p", 'rb'))
        allHouseKeepers = AHK.housekeepers
    except:
        allHouseKeepers = []
    try:
        covidRules = pickle.load(open("settings.p", 'rb'))
    except:
        covidRules = 0

    program = launchHKGUI(allHouseKeepers, covidRules)

    houseKeepers = program.activeHousekeepers
    allHouseKeepers = program.allHouseKeepers
    covidRules = program.covidRules
    #print(covidRules)

    if not houseKeepers:
        sys.exit()

    loadBar = Thread(target = loading)
    loadBar.daemon = True
    loadBar.start()


    ##################################################################
    #   USES TIKA LIBRARY TO EXTRACT TEXT FROM PDF
    ##################################################################
    #try:
    parsed = parser.from_file("../Custom Associates Report.pdf")
    rawText = parsed["content"]
    

    ##################################################################
    #   USING RAWTEXT, MANUALLY SPLITS BY LINES, IGNORING EMPTY LINES
    ##################################################################
    try:
        line = ""
        linedText = []
        for char in rawText:
            if char == '\n':
                if line: linedText.append(line)
                line = ""
            else:
                line += char
    except:
        generalError("There was an error communicating with a Java toolkit."
                     + " Please try restarting this computer to resolve the"
                     + " issue. If there is an update to Windows pending it is"
                     + " likely the culprit so please check for updates.")
        sys.exit()

    ##################################################################
    #   VERIFIES THAT CAR IS CURRENT BASED ON PRINTED DATETIME
    ##################################################################
    i = 0
    reasonable = False
    for line in linedText:
        if i > 5:
            break
        if "Printed:" in line:
            relevant = line.split()[1].split("/")
            dt = date(int(relevant[2]), int(relevant[0]), int(relevant[1]))
            if dt == today and "AM" in line:
                reasonable = True
    if not reasonable:
        stop = outdatedCARError()
        if stop:
            sys.exit()

    ##################################################################
    #   REMOVES ALL IRRELEVANT LINES LEAVING ONLY ROOM DATA AND TOTALS
    #   COMPARES NUMBER OF RELEVANT LINES COMPARED TO PDF DEFINED TOTALS
    ##################################################################
    i = 1
    relevantLines = []
    interestFlag = False
    for line in linedText:
        words = line.split()
        if words[0] in ["Custom", "Room"]: interestFlag = False
        if interestFlag:
            relevantLines.append(line)
            i+=1
        if words[-1] == "Remarks": interestFlag = True
    assert(i - 2 == int(relevantLines[-1][-3:]))


    ##################################################################
    #   GENERATES ONE ROOM OBJ PER LINE OF ROOM DATA
    ##################################################################
    dayList = ["mon", "tue", "wed", "thu", "fri", "sat", "sun", "monday",
           "mondays", "mons", "tuesdays", "tues", "tuesday", "wednesday",
           "weds", "wednesdays", "thursday", "thursdays", "thur", "thurs",
           "friday", "fridays", "fris", "saturday", "saturdays", "sats",
           "sunday", "sundays","suns"]
    try:
        roomTypes = ["KING","QQ","CK","CQ", "ACK", "ACQ", "SQ", "UF"]
        cleanTypes = ["Stay", "C/O", "Dirty"]
        rooms = []
        for line in relevantLines[:-1]:
            roomNumber = int(line[-3:])
            longStay = False
            arrival = False
            checkFT = False
            #if "No Service" in line: continue
            if "Dirty" in line:
                roomType = [x for x in roomTypes if x in line]
                if "Yes" in line:
                    arrival = True
                cleanType = ["Dirty"]
                checkInDate = "07/12/1996"
                room = Room(roomNumber, roomType[0], cleanType[0], arrival, longStay, checkInDate)
                rooms.append(room)
            else:
                roomType = [x for x in roomTypes if x in line]
                cleanType = [x for x in cleanTypes if x in line]
                yesCounter = line.count("Yes")
                daysDefined = []
                for day in dayList:
                    if day in line.lower():
                        #print(day)
                        daysDefined.append(day)
                        if day[0:3] == dayList[wday]:
                            #print(day)
                            longStay = True
                        if day[0:3] == dayList[wday-1]:
                            #print(day)
                            checkFT = True
                if not daysDefined and yesCounter == 1:
                    longStay = True
                elif yesCounter == 2:
                    arrival = True
                checkInDate = line.split()[0]
                room = Room(roomNumber, roomType[0], cleanType[0], arrival, longStay, checkInDate, checkFT)
                
                if not daysDefined:
                    room.checkBlueness()
                    room.checkOrangeness()
                if "No Service" in line or daysDefined:
                    if room.longStay: rooms.append(room)
                    if room.checkFT: rooms.append(room)
                else: 
                    rooms.append(room)
    except:
        generalError("There was an error interpreting the Custom Associates."
                     + " Report. Unfortunately, this means housekeeping will"
                     + " need to be assigned manually today. Please e-mail"
                     + " the CAR pdf as an attachment to:\n"
                     + "andresg8@uci.edu")
        sys.exit()

    ##################################################################
    #   Partition rooms as being either a towelTidy or a checkOut
    ##################################################################
    towelTidies = []
    checkOuts = []
    for room in rooms:
        if (room.cleanType == "Stay"):
            towelTidies.append(room)
        else:
            checkOuts.append(room)


    ##################################################################
    #   User input defines number of housekeepers working today
    #   here, we divide the TT's and C/O's as evenly as possible
    #   and follow the schema of min C/O's at the end and min TT's
    #   at the beginning. End and beg refer to top and bottom floors
    #   respectively. This is to allow for more C/O's to those closer
    #   to the laundry room and balancing that scale by assigining them
    #   less TT's.
    ##################################################################
    numHK = len(houseKeepers)
    roomsPerHK = []
    for i in range(numHK):
        roomsPerHK.append([])
        
    coPerHK = smoothDivision(len(checkOuts), numHK)
    coPerHK.reverse()
    roomCounter = 0
    if coPerHK:
        for hk in range(len(roomsPerHK)):
            for i in range(coPerHK[hk]):
                roomsPerHK[hk].append(checkOuts[roomCounter])
                roomCounter += 1
                if roomCounter > len(checkOuts):
                    print("Partitioning Error!")
                    break

    ttPerHK = smoothDivision(len(towelTidies), numHK)
    roomCounter = 0
    if ttPerHK:
        for hk in range(len(roomsPerHK)):
            for i in range(ttPerHK[hk]):
                roomsPerHK[hk].append(towelTidies[roomCounter])
                roomCounter += 1
                if roomCounter > len(towelTidies):
                    print("Partitioning Error!")
                    break


    ##################################################################
    #   Program eliminates ALL .xlsx files from the user accessible
    #   directory. Due to the threat this imposes to innocent bystander
    #   files, this function only executes if the directory name implies
    #   consent. The reason for this wipe is to ensure that on every run,
    #   the user is given the peace of mind that they will be left with
    #   ONLY the HK sheets relevant to the day they're assigning. 
    ##################################################################
    try:
        true_cwd = os.getcwd()
        os.chdir("../")
        if os.getcwd().endswith("HK Assigner"):
            cwd = os.getcwd()
            for file in os.listdir(cwd):
                if file.endswith(".xlsx"):
                    os.remove(file)
        os.chdir(true_cwd)
    except:
        generalError("There was an error when clearing the previous day's excel"
                     + " sheets. Make sure none of yesterdays HK excel sheets are"
                     + " open on any of the network's computers and try again.")
        sys.exit()

    yellowFill = PatternFill(start_color = '00FFFF00',
                                end_color = '00FFFF00',
                                fill_type = "solid")
    blueFill = PatternFill(start_color = '0037CBFF',
                                end_color = '0037CBFF',
                                fill_type = "solid")
    greenFill = PatternFill(start_color = '005EF600',
                                end_color = '005EF600',
                                fill_type = "solid")
    orangeFill = PatternFill(start_color = '00FFC000',
                                end_color = '00FFC000',
                                fill_type = "solid")
    noFill = PatternFill(start_color = '00FFFFFF',
                                end_color = '00FFFFFF',
                                fill_type = "solid")


    i = 0
    for hk in roomsPerHK:
        name = houseKeepers[i].name
        wb = load_workbook("HK Template.xlsx")
        ws = wb["Sheet1"]
        aggregateRoomScore = 0
        ws["A1"] = name
        r = 3
        space_added = False
        for room in hk:
            tipo = room.cleanType
            if room.cleanType not in["C/O", "Dirty"]: tipo = "TT"
            if tipo == "TT" and not space_added:
                r += 1
                space_added = True
            row = str(r)
            camas = room.roomType
            if room.roomType in ("CK", "ACK", "KING", "SQ"): camas += " - 1"
            if room.roomType in ("CQ", "ACQ", "QQ"): camas += " - 2"
            ws["A" + row] = camas
            ws["B" + row] = room.roomNumber
            printipo = "C/O"
            if tipo == "TT": printipo = "TT"
            ws["C" + row] = printipo

            if covidRules:
                if tipo in ["C/O"]:
                    ws["A" + row].fill = orangeFill
                    ws["B" + row].fill = orangeFill
                    ws["C" + row].fill = orangeFill
                elif tipo in ["Dirty"]:
                    ws["A" + row].fill = yellowFill
                    ws["B" + row].fill = yellowFill
                    ws["C" + row].fill = yellowFill
            else:
                if tipo in ["C/O", "Dirty"] and room.arrival:
                    ws["A" + row].fill = yellowFill
                    ws["B" + row].fill = yellowFill
                    ws["C" + row].fill = yellowFill
            if tipo == "TT" and room.longStay:
                ws["A" + row].fill = blueFill
                ws["B" + row].fill = blueFill
                ws["C" + row].fill = blueFill
            if tipo == "TT" and room.checkFT:
                ws["A" + row].fill = greenFill
                ws["B" + row].fill = greenFill
                ws["C" + row].fill = greenFill
            
            aggregateRoomScore += room.roomNumber
            r += 1
        new = True
        for j in range(len(allHouseKeepers)):
            if allHouseKeepers[j].name == name:
                allHouseKeepers[j] = HouseKeeper(name, today, aggregateRoomScore)
                new = False
        if new: allHouseKeepers.append(HouseKeeper(name, today, aggregateRoomScore))
        i += 1
        wb.save("../" + name + ".xlsx")
        wb.close()


    wb = load_workbook('Pre-Room Check List.xlsx')
    ws = wb["Sheet1"]

    letters = "BCDEFGHIJKL" 


    i = 0
    l = 0
    pageNum = 1
    for hk in roomsPerHK:
        name = houseKeepers[i].name
        for room in hk:
            if room.cleanType in ["C/O", "Dirty"]:
                ws[letters[l] + "2"] = room.roomNumber
                ws[letters[l] + "3"] = name[:3]
                if room.arrival:
                    ws[letters[l] + "2"].fill = yellowFill
                else:
                    ws[letters[l] + "2"].fill = noFill
                l += 1
                if (l > 10):
                    queuePrintJob(pageNum, ws, wb)
                    pageNum += 1
                    l = 0
        i += 1
        
    if l > 0:
        for i in range(l, 11):
            ws[letters[i] + "2"] = ""
            ws[letters[i] + "3"] = ""
            ws[letters[i] + "2"].fill = noFill
        queuePrintJob(pageNum, ws, wb)
    
    wb = load_workbook('Laundry Min.xlsx')
    wb.save('../Laundry Min.xlsx')
    wb.close()




    AHK = AllHouseKeepers(allHouseKeepers)
    pickle.dump(AHK, open("housekeepers.p", 'wb'))
    pickle.dump(covidRules, open("settings.p", 'wb'))
    print("done!")
    sys.exit()
